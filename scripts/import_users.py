"""
Import actual users from an XLSX file into the PIA Tracker database.

Usage:
    python scripts/import_users.py <path-to-xlsx>
    python scripts/import_users.py <path-to-xlsx> --dry-run
    python scripts/import_users.py <path-to-xlsx> --container pia-postgres --db-user pia --db-name pia

    # Run on your PC against the air-gapped VM (DB is rootful podman on the VM):
    python scripts/import_users.py officers.csv --ssh root@192.168.0.240 --runtime podman
    python scripts/import_users.py officers.csv --ssh root@192.168.0.240 --runtime podman --dry-run

    # Also set each user's password_hash eagerly (initial password = their HRMS id):
    python scripts/import_users.py officers.csv --ssh root@192.168.0.240 --runtime podman --hash-password
    # ...or a single shared initial password for everyone:
    python scripts/import_users.py officers.csv --ssh root@192.168.0.240 --runtime podman \
        --hash-password --password 'Welcome@123'

The script:
  1. Reads rows from the "Users" sheet (xlsx) or a CSV with equivalent headers.
  2. Validates each row against live designation/zone data from the DB.
  3. Skips rows where the email already exists.
  4. Inserts valid rows with is_demo = FALSE into the users table.
  5. Optionally (--hash-password) BCrypt-hashes a password into password_hash instead
     of leaving it NULL for lazy first-login init.
  6. Prints a per-row result table and a summary.
"""

import argparse
import csv
import subprocess
import sys
import uuid
from pathlib import Path
from typing import Optional


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Import users from XLSX into PIA Tracker DB.")
    p.add_argument("xlsx", help="Path to the filled-in XLSX or CSV file")
    p.add_argument("--dry-run", action="store_true", help="Validate only, no DB writes")
    p.add_argument("--container", default="pia-postgres", help="Container name (default: pia-postgres)")
    p.add_argument("--db-user",  default="pia",           help="PostgreSQL user (default: pia)")
    p.add_argument("--db-name",  default="pia",           help="PostgreSQL database (default: pia)")
    p.add_argument("--ssh", default=None, metavar="USER@HOST",
                   help="Reach the DB on a remote VM over SSH, e.g. root@192.168.0.240. "
                        "Use this to run the import from your PC — the container lives on the VM.")
    p.add_argument("--runtime", default="docker", choices=["docker", "podman"],
                   help="Container runtime holding the DB (default: docker; use 'podman' for the RHEL VM).")
    p.add_argument("--sudo", action="store_true",
                   help="Prefix the container command with sudo (needed for rootful podman on the VM).")
    p.add_argument("--hash-password", action="store_true",
                   help="Compute a BCrypt hash and store it in password_hash at insert time, "
                        "instead of leaving it NULL for lazy first-login init. "
                        "By default each user's HRMS id is hashed (so the initial password stays "
                        "the HRMS id); pass --password to hash a fixed value for everyone instead.")
    p.add_argument("--password", default=None, metavar="PLAINTEXT",
                   help="With --hash-password, the plaintext to hash for EVERY imported user. "
                        "Omit to hash each user's own HRMS id.")
    p.add_argument("--out-sql", default=None, metavar="FILE",
                   help="Offline: don't touch the DB — write idempotent INSERT statements to FILE "
                        "(with BCrypt hashes if --hash-password). scp it to the VM and apply with "
                        "psql -f. No --ssh / DB connection needed.")
    return p.parse_args()


def make_bcrypt_hasher():
    """Return a fn raw->hash matching Spring's BCryptPasswordEncoder() (strength 10, $2a$)."""
    try:
        import bcrypt
    except ImportError:
        sys.exit("--hash-password needs the 'bcrypt' package on this machine: pip install bcrypt")

    def _hash(raw: str) -> str:
        # $2a$ prefix + rounds=10 to match the backend encoder exactly; Spring also
        # accepts $2b/$2y, but matching avoids any doubt.
        return bcrypt.hashpw(raw.encode("utf-8"), bcrypt.gensalt(rounds=10, prefix=b"2a")).decode("ascii")

    return _hash


# ── DB helpers ─────────────────────────────────────────────────────────────────

def build_psql_cmd(args) -> list:
    """
    Assemble the argv that runs psql against the target DB. Three optional layers:
      --ssh USER@HOST  → run on the remote VM (the container lives there, not locally)
      --sudo           → rootful podman on the VM needs sudo
      --runtime podman → podman instead of docker
    SQL is fed on stdin (not `-c`) so it survives SSH transport without any quoting.
    """
    cmd = []
    if args.ssh:
        cmd += ["ssh", args.ssh]
    if args.sudo:
        cmd += ["sudo"]
    cmd += [args.runtime, "exec", "-i", args.container,
            "psql", "-U", args.db_user, "-d", args.db_name, "-t", "-A"]
    return cmd


def psql(base_cmd: list, sql: str) -> str:
    result = subprocess.run(base_cmd, input=sql, capture_output=True, text=True)
    if result.returncode != 0:
        sys.exit(f"DB error: {result.stderr.strip()}")
    return result.stdout.strip()


def fetch_set(base_cmd, sql) -> set:
    out = psql(base_cmd, sql)
    return {line.strip() for line in out.splitlines() if line.strip()}


# ── Designation alias map ──────────────────────────────────────────────────────
# Maps external / HR-system labels to system designation codes.
# Keys are normalised to uppercase with extra spaces collapsed.

# Coarse family map. Designation codes drive role resolution via
# designation_default_roles, so every officer must land on the *_C family code
# that carries the operational role (not the granular sub-codes, which only
# carry ROLE_APPROVER_GENERIC):
#   DY CE*  -> DY_CE_C  (ROLE_DY_CE_C)
#   CE*     -> CE_C     (ROLE_CE_C)
#   CAO*    -> CAO_C    (ROLE_CAO_C)
#   ED*     -> EDGS_CI  (ROLE_EDGS_CI)
# The prefix rules in resolve_designation() do the heavy lifting; this table is
# only for exact one-offs that the prefixes would otherwise misclassify.
DESIGNATION_ALIASES = {
    "EDGS/C-I": "EDGS_CI",
}

def resolve_designation(raw: str) -> tuple[str, Optional[str]]:
    """
    Returns (system_code, alias_note).
    alias_note is set when an alias was used so the caller can log it.
    """
    import re
    normalised = re.sub(r"\s+", " ", raw.strip().upper())
    if normalised in DESIGNATION_ALIASES:
        code = DESIGNATION_ALIASES[normalised]
        return code, f"mapped from '{raw}'"

    # Coarse prefix collapse to the role-bearing family code. Compact the label to
    # letters only first, so 'DY CE', 'DY. CE', 'DyCE', 'DY-CE/Bridgeline/HQ' all read
    # the same. Order matters: "DYCE" must be tested before "CE"; "ED"/Executive
    # Director before "CE".
    compact = re.sub(r"[^A-Z]", "", normalised)
    if compact.startswith("DYCE"):
        return "DY_CE_C", f"mapped from '{raw}'"
    if compact.startswith("CAO"):
        return "CAO_C", f"mapped from '{raw}'"
    if compact.startswith("ED") or "EXECUTIVEDIRECTOR" in compact:
        return "EDGS_CI", f"mapped from '{raw}'"
    if compact.startswith("CE"):
        return "CE_C", f"mapped from '{raw}'"

    return raw.strip(), None


# ── Row loaders (XLSX + CSV) ─────────────────────────────────────────────────────
# Both return a list of dicts: {name, email, desig_raw, zone, employee_id}.

# Header aliases → canonical field. Matching is case-insensitive with spaces/
# underscores/punctuation stripped, so "Employee ID", "employee_id", "EmpId" all match.
HEADER_ALIASES = {
    "name": "name", "fullname": "name", "officername": "name",
    "employeename": "name", "empname": "name",
    "email": "email", "emailid": "email", "mail": "email",
    # Prefer the human-readable designation text; the numeric desig_code is ignored.
    "designation": "desig_raw", "designationcode": "desig_raw", "desig": "desig_raw",
    "desigdesc": "desig_raw", "designationdesc": "desig_raw",
    "zone": "zone", "zonecode": "zone", "railway": "zone",
    "employeeid": "employee_id", "empid": "employee_id", "pfno": "employee_id",
    "employeeno": "employee_id", "hrmsid": "employee_id", "emphrmsid": "employee_id",
}


def _norm_header(h: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]", "", (h or "").strip().lower())


def load_csv_rows(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return []
        # Map each column index to a canonical field name via the alias table.
        col_map: dict[int, str] = {}
        for i, h in enumerate(header):
            field = HEADER_ALIASES.get(_norm_header(h))
            if field:
                col_map[i] = field
        present = set(col_map.values())
        missing = {"name", "desig_raw"} - present
        if missing:
            sys.exit(
                f"CSV header is missing required column(s): {', '.join(sorted(missing))}.\n"
                f"Found headers: {header}\n"
                f"Expected something matching: Name, Designation (+ Zone, Employee ID)."
            )
        # Email may be absent — it is generated from employee_id + zone (see main()).
        if "email" not in present and not {"employee_id", "zone"} <= present:
            sys.exit(
                "CSV has no Email column, and can't generate one: need both an "
                "Employee ID (hrms id) column and a Zone column to synthesise "
                "{hrms_id}@{zone}.railnet.gov.in.\n"
                f"Found headers: {header}"
            )
        rows = []
        for values in reader:
            rec = {"name": "", "email": "", "desig_raw": "", "zone": None, "employee_id": None}
            for i, field in col_map.items():
                if i < len(values):
                    rec[field] = (values[i] or "").strip()
            if not any([rec["name"], rec["email"], rec["desig_raw"]]):
                continue
            rows.append(rec)
        return rows


def load_xlsx_rows(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required for .xlsx files: pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Users" not in wb.sheetnames:
        sys.exit('XLSX must contain a sheet named "Users".')
    ws = wb["Users"]
    rows = []
    # Row 3 = headers, row 4 = sample — data starts at row 5
    for row_num in range(5, ws.max_row + 1):
        def cell(col): return ws.cell(row=row_num, column=col).value
        name  = str(cell(1)).strip() if cell(1) else ""
        email = str(cell(2)).strip() if cell(2) else ""
        desig = str(cell(3)).strip() if cell(3) else ""
        zone  = str(cell(4)).strip() if cell(4) else None
        emp   = str(cell(5)).strip() if cell(5) else None
        if not any([name, email, desig]):
            continue
        rows.append({"name": name, "email": email, "desig_raw": desig,
                     "zone": zone or None, "employee_id": emp or None})
    return rows


def load_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        return load_csv_rows(path)
    return load_xlsx_rows(path)


# ── Offline SQL emitter ──────────────────────────────────────────────────────

def write_sql_file(path: str, valid_rows: list, hasher, fixed_password) -> int:
    """
    Write idempotent INSERT statements (one per user) to `path`. Mirrors the live-insert
    columns; adds password_hash/password_updated_at when a hasher is given. Each row is
    ON CONFLICT (employee_id) DO NOTHING so re-applying is safe, and each is its own
    statement so a bad row (e.g. unknown designation) is reported by psql without aborting
    the rest. Returns the number of statements written.
    """
    out = [
        "-- PIA Tracker — bulk officer import (generated by scripts/import_users.py).",
        "-- Idempotent: ON CONFLICT (employee_id) DO NOTHING. Apply with:",
        "--   sudo podman exec -i pia-postgres psql -U pia -d pia -f /tmp/users.sql",
        "",
    ]
    for r in valid_rows:
        uid       = str(uuid.uuid4())
        name_esc  = r["name"].replace("'", "''")
        email_esc = r["email"].replace("'", "''")
        desig_esc = r["designation_code"].replace("'", "''")
        emp       = r["employee_id"]
        emp_val   = f"'{emp.replace(chr(39), chr(39)*2)}'" if emp else "NULL"
        zone_val  = (f"(SELECT id FROM zones WHERE code = '{r['zone_code']}')"
                     if r["zone_code"] else "NULL")

        pw_cols = ""
        pw_vals = ""
        if hasher is not None:
            raw_pw = fixed_password if fixed_password else (emp or "")
            if raw_pw:
                pw_hash = hasher(raw_pw).replace("'", "''")
                pw_cols = ", password_hash, password_updated_at"
                pw_vals = f", '{pw_hash}', now()"

        out.append(
            "INSERT INTO users (id, name, email, designation_code, primary_zone_id, employee_id, "
            f"is_demo, is_active, is_deleted, is_system_user{pw_cols}, created_at, updated_at) "
            f"VALUES ('{uid}', '{name_esc}', '{email_esc}', '{desig_esc}', {zone_val}, {emp_val}, "
            f"FALSE, TRUE, FALSE, FALSE{pw_vals}, now(), now()) "
            "ON CONFLICT (employee_id) DO NOTHING;"
        )

    Path(path).write_text("\n".join(out) + "\n", encoding="utf-8")
    return len(valid_rows)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    container = args.container
    db_name   = args.db_name
    offline   = bool(args.out_sql)
    base_cmd  = None

    # ── Load reference data from DB (skipped in offline --out-sql mode) ────────
    if offline:
        # No DB: validation against live designations/zones/emails is not possible,
        # so the generated SQL relies on FK constraints + ON CONFLICT at apply time.
        valid_designations = valid_zones = existing_emails = None
        print("Offline mode: generating a .sql file (no DB connection).\n")
    else:
        base_cmd = build_psql_cmd(args)
        via = f"via ssh {args.ssh} " if args.ssh else ""
        print(f"Connecting to DB {via}({container} / {db_name})…")
        valid_designations = fetch_set(base_cmd, "SELECT code FROM designations;")
        valid_zones = fetch_set(base_cmd, "SELECT code FROM zones WHERE is_active = true;")
        existing_emails = fetch_set(base_cmd, "SELECT lower(email) FROM users WHERE is_deleted = false;")
        print(f"  {len(valid_designations)} designations, {len(valid_zones)} zones, "
              f"{len(existing_emails)} existing users loaded.\n")

    # ── Read input file (XLSX or CSV) ─────────────────────────────────────────
    raw_rows = load_rows(xlsx_path)

    rows = []
    for idx, rec in enumerate(raw_rows, start=1):
        name        = rec["name"]
        email       = rec["email"]
        desig_raw   = rec["desig_raw"]
        zone        = rec["zone"]
        employee_id = rec["employee_id"]
        row_num     = idx  # 1-based data row for display

        # Skip the italic sample row if user left it in
        if email == "rajesh.sharma@nr.railnet.gov.in":
            continue

        # Generate email when the source has none: {hrms_id}@{zone}.railnet.gov.in
        # (matches the existing seeded users; hrms_id is unique so email stays unique).
        if not email and employee_id and zone:
            email = f"{employee_id.lower()}@{zone.lower()}.railnet.gov.in"

        desig, alias_note = resolve_designation(desig_raw)

        errors = []
        notes  = []
        if alias_note:
            notes.append(alias_note)
        # Railway Board officers (Railway/Zone == "RB") always land on the ED family
        # (EDGS_CI), regardless of their designation text.
        if (zone or "").strip().upper() == "RB":
            if desig != "EDGS_CI":
                notes.append(f"Railway 'RB' -> designation EDGS_CI (was {desig})")
            desig = "EDGS_CI"
        if not name:
            errors.append("Name is required")
        if not email:
            errors.append("Email is required")
        elif "@" not in email:
            errors.append("Email looks invalid")
        if not desig_raw:
            errors.append("Designation Code is required")
        elif valid_designations is not None and desig not in valid_designations:
            errors.append(f"Unknown designation '{desig_raw}'")
        if valid_zones is not None and zone and zone not in valid_zones:
            errors.append(f"Unknown zone '{zone}'")
        if existing_emails is not None and email and email.lower() in existing_emails:
            errors.append("Email already exists in DB -- skipped")

        rows.append({
            "row_num": row_num,
            "name": name,
            "email": email,
            "designation_code": desig,
            "zone_code": zone or None,
            "employee_id": employee_id or None,
            "errors": errors,
            "notes": notes,
        })

    if not rows:
        print("No data rows found in the XLSX (rows 5+). Nothing to import.")
        return

    # ── Print validation results ──────────────────────────────────────────────
    print(f"{'Row':<5} {'Status':<10} {'Name':<25} {'Email':<40} {'Desig':<15} {'Zone':<6}  Notes")
    print("-" * 120)

    valid_rows   = []
    skipped_rows = []

    for r in rows:
        ok = not r["errors"]
        status = "OK" if ok else "SKIP"
        all_notes = r["notes"] + r["errors"]
        note_str  = "; ".join(all_notes) if all_notes else ""
        print(f"{r['row_num']:<5} {status:<10} {r['name'][:24]:<25} {r['email'][:39]:<40} "
              f"{r['designation_code'][:14]:<15} {(r['zone_code'] or ''):<6}  {note_str}")
        if ok:
            valid_rows.append(r)
        else:
            skipped_rows.append(r)

    print("-" * 120)
    print(f"\n{len(valid_rows)} valid  |  {len(skipped_rows)} skipped\n")

    if not valid_rows:
        print("Nothing to import.")
        return

    hasher = make_bcrypt_hasher() if args.hash_password else None
    if hasher is not None:
        which = "the fixed --password value" if args.password else "each user's HRMS id"
        print(f"Password mode: BCrypt-hashing {which} into password_hash.")

    # ── Offline: write a .sql file to scp + apply on the VM ────────────────────
    if offline:
        n = write_sql_file(args.out_sql, valid_rows, hasher, args.password)
        print(f"\nWrote {n} INSERT statement(s) to {args.out_sql}")
        print("Apply it on the VM (one scp + one psql, no per-row prompts):")
        print(f"  scp {args.out_sql} root@<VM>:/tmp/users.sql")
        print("  ssh root@<VM> \"sudo podman cp /tmp/users.sql pia-postgres:/tmp/users.sql && "
              "sudo podman exec pia-postgres psql -U pia -d pia -f /tmp/users.sql\"")
        return

    if args.dry_run:
        print("Dry run — no changes written to DB.")
        return

    # ── Insert ────────────────────────────────────────────────────────────────
    print(f"Inserting {len(valid_rows)} user(s)…")

    inserted = 0
    failed   = 0

    for r in valid_rows:
        uid         = str(uuid.uuid4())
        name_esc    = r["name"].replace("'", "''")
        email_esc   = r["email"].replace("'", "''")
        desig_esc   = r["designation_code"].replace("'", "''")
        emp_id_val  = f"'{r['employee_id'].replace(chr(39), chr(39)*2)}'" if r["employee_id"] else "NULL"

        # Resolve zone_id from zone code
        if r["zone_code"]:
            zone_id_val = f"(SELECT id FROM zones WHERE code = '{r['zone_code']}')"
        else:
            zone_id_val = "NULL"

        # Optional eager password hash. Default source is the user's HRMS id, so the
        # initial password matches the app's convention — just pre-set, not lazy.
        pw_cols = ""
        pw_vals = ""
        if hasher is not None:
            raw_pw = args.password if args.password else (r["employee_id"] or "")
            if raw_pw:
                pw_hash = hasher(raw_pw).replace("'", "''")
                pw_cols = ", password_hash, password_updated_at"
                pw_vals = f", '{pw_hash}', now()"

        sql = f"""
INSERT INTO users (id, name, email, designation_code, primary_zone_id, employee_id,
                   is_demo, is_active, is_deleted, is_system_user{pw_cols},
                   created_at, updated_at)
VALUES (
    '{uid}', '{name_esc}', '{email_esc}', '{desig_esc}', {zone_id_val}, {emp_id_val},
    FALSE, TRUE, FALSE, FALSE{pw_vals},
    now(), now()
);
""".strip()

        out = psql(base_cmd, sql)
        if "INSERT 0 1" in out or out == "INSERT 0 1" or out == "":
            # psql -t -A strips the result; check via a count query
            count = psql(base_cmd, f"SELECT count(*) FROM users WHERE email = '{email_esc}';")
            if count == "1":
                print(f"  OK Inserted: {r['name']} <{r['email']}>")
                inserted += 1
            else:
                print(f"  FAIL Failed: {r['name']} <{r['email']}> -- {out}")
                failed += 1
        else:
            print(f"  FAIL Failed: {r['name']} <{r['email']}> -- {out}")
            failed += 1

    print(f"\nDone. {inserted} inserted, {failed} failed, {len(skipped_rows)} skipped.\n")
    if inserted > 0:
        print("Users are marked is_demo=FALSE and will appear in the 'Users' section on the login page.")
        if hasher is not None:
            src = "the shared --password value" if args.password else "each user's HRMS id"
            print(f"password_hash was pre-set (BCrypt of {src}); no first-login init needed.")
        else:
            print("password_hash left NULL — initial password is each user's HRMS id (set on first login).")


if __name__ == "__main__":
    main()

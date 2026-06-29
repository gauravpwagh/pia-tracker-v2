"""Generates the user-import XLSX template at scripts/user_import_template.xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

OUT = Path(__file__).parent / "user_import_template.xlsx"

DESIGNATIONS = [
    ("EDGS_CI",       "EDGS/C-I"),
    ("CAO_C",         "CAO/C"),
    ("CE_C",          "CE/C"),
    ("CE_PLANNING",   "CE/Planning"),
    ("DY_CE_C",       "Dy CE/C"),
    ("DY_CE_PLANNING","Dy CE/Planning"),
    ("DY_CE_DESIGN",  "Dy CE/Design"),
    ("DY_CE",         "Dy CE"),
    ("SR_DEN",        "Sr DEN"),
    ("SR_DEN_CO",     "Sr DEN/Co"),
    ("CBE",           "CBE"),
    ("DY_CE_BRIDGE",  "Dy CE/Bridge"),
    ("CTE",           "CTE"),
    ("DY_CE_TRACK",   "Dy CE/Track"),
    ("CPDE",          "CPDE"),
    ("PCE",           "PCE"),
    ("DY_CSTE",       "Dy CSTE"),
    ("SR_DSTE",       "Sr DSTE"),
    ("CSTE_CON",      "CSTE/Con"),
    ("CSTE_OL",       "CSTE/OL"),
    ("PSCTE",         "PSCTE"),
    ("DY_CEE",        "Dy CEE"),
    ("SR_DEE_TRD",    "Sr DEE/TRD"),
    ("CEE_CON",       "CEE/Con"),
    ("PCEE",          "PCEE"),
    ("SR_DOM",        "Sr DOM"),
    ("PCOM",          "PCOM"),
    ("SR_DCM",        "Sr DCM"),
    ("ADRM",          "ADRM"),
    ("DRM",           "DRM"),
    ("CTPM",          "CTPM"),
    ("PCSO",          "PCSO"),
    ("CRS",           "CRS"),
    ("GM",            "GM"),
    ("ADMIN",         "Admin"),
    ("SUPER_ADMIN",   "SAdmin"),
]

ZONES = [
    ("CR",   "Central Railway"),
    ("ECR",  "East Central Railway"),
    ("ECOR", "East Coast Railway"),
    ("ER",   "Eastern Railway"),
    ("KR",   "Konkan Railway"),
    ("NCR",  "North Central Railway"),
    ("NER",  "North Eastern Railway"),
    ("NFR",  "Northeast Frontier Railway"),
    ("NR",   "Northern Railway"),
    ("NWR",  "North Western Railway"),
    ("SCR",  "South Central Railway"),
    ("SECR", "South East Central Railway"),
    ("SER",  "South Eastern Railway"),
    ("SR",   "Southern Railway"),
    ("SWR",  "South Western Railway"),
    ("WCR",  "West Central Railway"),
    ("WR",   "Western Railway"),
]

wb = openpyxl.Workbook()

# ── Sheet 1: Users ─────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Users"

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHDR_FILL   = PatternFill("solid", fgColor="D6E4F0")
SUBHDR_FONT   = Font(name="Arial", bold=True, color="1F4E79", size=9)
CELL_FONT     = Font(name="Arial", size=10)
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
OPTIONAL_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN          = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Title row
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "PIA Tracker — User Import Template"
title_cell.font  = Font(name="Arial", bold=True, size=13, color="1F4E79")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Sub-header note
ws.merge_cells("A2:F2")
note = ws["A2"]
note.value = "Fill in the yellow columns (required). Grey columns are optional. Do not modify column headers."
note.font  = Font(name="Arial", italic=True, size=9, color="595959")
note.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

# Column headers (row 3)
COLS = [
    ("A", "Name *",           True,  28),
    ("B", "Email *",          True,  32),
    ("C", "Designation Code *", True, 20),
    ("D", "Zone Code",        False, 14),
    ("E", "Employee ID",      False, 16),
    ("F", "Notes",            False, 30),
]

for col_letter, label, required, width in COLS:
    cell = ws[f"{col_letter}3"]
    cell.value     = label
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[3].height = 22

# Sample row
sample = ["Rajesh Sharma", "rajesh.sharma@nr.railnet.gov.in", "DY_CE_C", "NR", "NR12345", ""]
for i, val in enumerate(sample, 1):
    cell = ws.cell(row=4, column=i)
    cell.value     = val
    cell.font      = Font(name="Arial", size=10, color="595959", italic=True)
    cell.fill      = PatternFill("solid", fgColor="F7F7F7")
    cell.border    = BORDER
    cell.alignment = Alignment(vertical="center")

# Data rows 5–204 (200 rows)
for row in range(5, 205):
    for col_idx, (col_letter, _, required, _) in enumerate(COLS, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font      = CELL_FONT
        cell.fill      = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.border    = BORDER
        cell.alignment = Alignment(vertical="center")

# Freeze header rows
ws.freeze_panes = "A4"

# Data validation — Designation Code dropdown
desig_codes = ",".join(d[0] for d in DESIGNATIONS)
dv_desig = DataValidation(
    type="list",
    formula1=f'"{ desig_codes }"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Designation",
    error="Choose a code from the dropdown or see the Reference sheet.",
)
ws.add_data_validation(dv_desig)
dv_desig.sqref = "C5:C204"

# Data validation — Zone Code dropdown
zone_codes = ",".join(z[0] for z in ZONES)
dv_zone = DataValidation(
    type="list",
    formula1=f'"{ zone_codes }"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid Zone",
    error="Choose a zone code from the dropdown or see the Reference sheet.",
)
ws.add_data_validation(dv_zone)
dv_zone.sqref = "D5:D204"

# Row 4 label
ws["A4"].font = Font(name="Arial", size=10, color="595959", italic=True)

# ── Sheet 2: Reference ─────────────────────────────────────────────────────────
ref = wb.create_sheet("Reference")
ref.sheet_state = "visible"

ref_title_fill = PatternFill("solid", fgColor="1F4E79")
ref_title_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ref_cell_font  = Font(name="Arial", size=10)

# Designation table
ref.merge_cells("A1:B1")
ref["A1"].value = "Designation Codes"
ref["A1"].font  = ref_title_font
ref["A1"].fill  = ref_title_fill
ref["A1"].alignment = Alignment(horizontal="center")
ref["B1"].fill  = ref_title_fill

ref["A2"].value = "Code";          ref["A2"].font = Font(name="Arial", bold=True, size=10)
ref["B2"].value = "Short Label";   ref["B2"].font = Font(name="Arial", bold=True, size=10)

for i, (code, label) in enumerate(DESIGNATIONS, 3):
    ref.cell(row=i, column=1).value = code;  ref.cell(row=i, column=1).font = ref_cell_font
    ref.cell(row=i, column=2).value = label; ref.cell(row=i, column=2).font = ref_cell_font

ref.column_dimensions["A"].width = 20
ref.column_dimensions["B"].width = 18

# Zone table (column D–E)
ref.merge_cells("D1:E1")
ref["D1"].value = "Zone Codes"
ref["D1"].font  = ref_title_font
ref["D1"].fill  = ref_title_fill
ref["D1"].alignment = Alignment(horizontal="center")
ref["E1"].fill  = ref_title_fill

ref["D2"].value = "Code";   ref["D2"].font = Font(name="Arial", bold=True, size=10)
ref["E2"].value = "Zone";   ref["E2"].font = Font(name="Arial", bold=True, size=10)

for i, (code, name) in enumerate(ZONES, 3):
    ref.cell(row=i, column=4).value = code; ref.cell(row=i, column=4).font = ref_cell_font
    ref.cell(row=i, column=5).value = name; ref.cell(row=i, column=5).font = ref_cell_font

ref.column_dimensions["D"].width = 8
ref.column_dimensions["E"].width = 28

# Notes table (column G)
ref["G1"].value = "Field Notes"
ref["G1"].font  = ref_title_font
ref["G1"].fill  = ref_title_fill
ref["G1"].alignment = Alignment(horizontal="center")
ref.column_dimensions["G"].width = 55

notes = [
    ("Name *",             "Full name, e.g. 'Rajesh Kumar Singh'"),
    ("Email *",            "Official railnet email — must be unique"),
    ("Designation Code *", "Must match a code from the Designation Codes table"),
    ("Zone Code",          "Leave blank for system/admin users (ADMIN, SUPER_ADMIN)"),
    ("Employee ID",        "Optional HRMS / PF number. Leave blank if unknown."),
    ("Notes",              "Any free-text notes — not imported into the database"),
]
ref["G2"].value = "Field"; ref["G2"].font = Font(name="Arial", bold=True, size=10)
ref["H2"].value = "Description"; ref["H2"].font = Font(name="Arial", bold=True, size=10)
ref.column_dimensions["H"].width = 55

for i, (field, desc) in enumerate(notes, 3):
    ref.cell(row=i, column=7).value  = field; ref.cell(row=i, column=7).font = ref_cell_font
    ref.cell(row=i, column=8).value  = desc;  ref.cell(row=i, column=8).font = ref_cell_font

wb.save(OUT)
print(f"Template written to {OUT}")
